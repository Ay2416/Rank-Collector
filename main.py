# Discord bot import
import discord
from discord import app_commands
from discord import ui
from discord.ext import tasks
import os
from dotenv import load_dotenv
import glob
import ndjson
import datetime
import dateutil.parser
from mk8dx import Track
import asyncio
import math
import shutil
import openpyxl


# Bot start

# envの読込
load_dotenv()

# 変数宣言
# データ保存をシーズンごとに分ける用
now_season = 9
# add,deleteの操作がない時に処理をとめる時間（秒）
stop_time = 600
# 文字列にした時間を計算できる形式に変換する時に使用
time_format = '%Y-%m-%d %H:%M:%S.%f'
# 時間をJSTの時間に変換するときに使用
JST = datetime.timezone(datetime.timedelta(hours=+9), 'JST')
# 時間をUTCの時間に変換するときに使用
UTC = datetime.timezone(datetime.timedelta(hours=0), 'UTC')
# 頭に記号がある他のBotのコマンドに反応しないように登録する場所
prefixes = ['!', '#', '$', '%', '&', '(', ')', '=', '-', '~', '^', '|', '`', ':', '*', '+', ';', '<', ',', '>', '.', '?', '_']

intents = discord.Intents.all()
client = discord.Client(intents=intents)
tree = app_commands.CommandTree(client)

# Bot開始時にさせる処理
@client.event
async def on_ready():
    print("接続しました！")
    await client.change_presence(activity=discord.Game(name="rank collect"))

    await tree.sync()#スラッシュコマンドを同期
    print("グローバルコマンド同期完了！")

    # guild_jsonフォルダがあるかの確認
    files = glob.glob('./*')
    judge = 0

    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if(os.path.split(files[i])[1] == "guild_json"):
            print("guild_jsonファイルを確認しました！")
            judge = 1
            break

    if judge != 1:
        os.mkdir('guild_json')
        print("guild_jsonファイルがなかったため作成しました！")

    # delete_jsonフォルダがあるかの確認
    files = glob.glob('./*')
    judge = 0

    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if(os.path.split(files[i])[1] == "delete_json"):
            print("delete_jsonファイルを確認しました！")
            judge = 1
            break

    if judge != 1:
        os.mkdir('delete_json')
        print("delete_jsonファイルがなかったため作成しました！")
    
    # add_jsonフォルダがあるかの確認
    files = glob.glob('./*')
    judge = 0

    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if(os.path.split(files[i])[1] == "add_json"):
            print("add_jsonファイルを確認しました！")
            judge = 1
            break

    if judge != 1:
        os.mkdir('add_json')
        print("add_jsonファイルがなかったため作成しました！")
        
    # user_jsonフォルダがあるかの確認
    files = glob.glob('./*')
    judge = 0

    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if(os.path.split(files[i])[1] == "user_json"):
            print("user_jsonファイルを確認しました！")
            judge = 1
            break

    if judge != 1:
        os.mkdir('user_json')
        print("user_jsonファイルがなかったため作成しました！")
    
    # language_jsonフォルダがあるかの確認
    judge = 0
    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if(os.path.split(files[i])[1] == "language_json"):
            print("language_jsonファイルを確認しました！")
            judge = 1
            break

    if judge != 1:
        os.mkdir('language_json')
        print("language_jsonファイルがなかったため作成しました！")
    
    # 定期的に動かすループ処理の開始
    time_check.start()
    print("時間の確認を開始します！")

# サーバーに招待された場合に特定の処理をする
@client.event
async def on_guild_join(guild):
    file = str(guild.id) + ".ndjson"

    content = {
        "language_mode" : "ja"
    }

    with open('./language_json/' + file, 'a') as f:
        writer = ndjson.writer(f)
        writer.writerow(content)
    
    print("招待されたため" + str(guild.id) + "のlanguage jsonを作成しました。")

# サーバーからキック、BANされた場合に特定の処理をする
@client.event
async def on_guild_remove(guild):
    file = str(guild.id) + ".ndjson"
    os.remove("./language_json/" + file)

    print("キックまたはBANされたため、" + str(guild.id) + "のlanguage jsonを削除しました。")
    
    files = glob.glob('./guild_json/*.ndjson')
    judge = 0

    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if os.path.split(files[i])[1] == str(guild.id) + ".ndjson":
            judge = 1
            break
    
    if judge == 1:
        os.remove("./guild_json/" + file)
        print("キックまたはBANされたため、" + str(guild.id) + "のguild jsonを削除しました。")
    
    files = glob.glob('./add_json/*')
    judge = 0

    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if os.path.split(files[i])[1] == str(guild.id):
            judge = 1
            break
    
    if judge == 1:
        os.remove("./add_json/" + file)
        print("キックまたはBANされたため、" + str(guild.id) + "のadd jsonを削除しました。")

    files = glob.glob("./delete_json/" + str(guild.id) + "/")
    judge = 0

    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if os.path.split(files[i])[1] == str(guild.id):
            judge = 1
            break
    
    if judge == 1:
        os.remove("./delete_json/" + str(guild.id) + "/" )
        print("キックまたはBANされたため、" + str(guild.id) + "のdelete jsonを削除しました。")

#メッセージを取得した時に実行される
@client.event
async def on_message(message):

    # Botのメッセージは除外
    if message.author.bot:
        return
    
    # 頭に記号がある場合は除外
    for i in range(0, len(prefixes)):
        if message.content.startswith(prefixes[i]) == True:
            return

    try:
        # 言語の確認
        file = str(message.guild.id) + ".ndjson"

        with open('./language_json/' + file) as f:
            read_data = ndjson.load(f)

        language = read_data[0]["language_mode"]

        # 現在のモードは何かを確認
        file = str(message.guild.id) + ".ndjson"

        #print(message.guild.id)

        # 同じチャンネルのデータがある部分を探す
        # guild_jsonの読込
        with open('./guild_json/' + file) as f:
            read_data = ndjson.load(f)
        
        data_location = 0
        for i in range(0, len(read_data)):
            if read_data[i]["channel"] == message.channel.id:
                data_location = i
                #print(read_data)
                break

        mode = read_data[data_location]["mode"]

        # コマンドが送信されたチャンネルでしか反応しないようにする
        if message.channel.id == read_data[data_location]["channel"]:

            # モードがaddの場合
            if mode == "add":
                # 入力されたメッセージを取得
                #print(message.channel.id)
                msg = message.content.split()
                course = str(msg[0])
                rank = int(msg[1])

                # 入力されたコース名がデータにあるかの判定
                try:
                    name = Track.from_nick(course)
                    print(name.abbr)
                except Exception as e:
                    if language == "ja":
                        embed=discord.Embed(title="そのコースデータは存在しません！", color=0xff0000)
                        embed.add_field(name="入力されたコース名を確認してください。", value="", inline=False)
                    elif language == "en":
                        embed=discord.Embed(title="Not course data!", color=0xff0000)
                        embed.add_field(name="Check typing course name.", value="", inline=False)
                    await message.channel.send(embed=embed)

                    # guild_jsonのtimeを更新
                    # 同じチャンネルのデータがある部分を探す
                    # guild_jsonの読込
                    with open('./guild_json/' + str(message.guild.id) + '.ndjson') as f:
                        read_data = ndjson.load(f)
                    
                    data_location1 = 0
                    for i in range(0, len(read_data)):
                        if read_data[i]["channel"] == message.channel.id:
                            data_location1 = i
                            #print(read_data)
                            break
                    
                    # ndjsonにtimeを書き込み
                    read_data[data_location1]["time"] = str(datetime.datetime.now())

                    os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                    for i in range(0, len(read_data)):
                        with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                            writer = ndjson.writer(f)
                            writer.writerow(read_data[i])
                    
                    return
                else:
                    #入力された順位が範囲内であるかの判定
                    if rank < 1 or rank > 12:
                        if language == "ja":
                            embed=discord.Embed(title="入力された順位が有効範囲外です！", color=0xff0000)
                            embed.add_field(name="有効な順位の範囲は1~12位です。", value="", inline=False)
                        elif language == "en":
                            embed=discord.Embed(title="The entered rank is out of the valid range!", color=0xff0000)
                            embed.add_field(name="The valid ranking range is 1~12.", value="", inline=False)
                               
                        await message.channel.send(embed=embed)

                        # guild_jsonのtimeを更新
                        # 同じチャンネルのデータがある部分を探す
                        # guild_jsonの読込
                        with open('./guild_json/' + str(message.guild.id) + '.ndjson') as f:
                            read_data = ndjson.load(f)
                        
                        data_location1 = 0
                        for i in range(0, len(read_data)):
                            if read_data[i]["channel"] == message.channel.id:
                                data_location1 = i
                                #print(read_data)
                                break
                        
                        # ndjsonにtimeを書き込み
                        read_data[data_location1]["time"] = str(datetime.datetime.now())

                        os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                        for i in range(0, len(read_data)):
                            with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                                writer = ndjson.writer(f)
                                writer.writerow(read_data[i])
                        
                        return
                    
                    # 既に登録があるかの確認
                    # コースを保存するユーザーファイル
                    files = glob.glob('./user_json/*')
                    judge = 0

                    for i in range(0, len(files)):
                        #print(os.path.split(files[i])[1])
                        if(os.path.split(files[i])[1] == str(message.author.id)):
                            judge = 1
                            break
                    
                    # なければ作成
                    if judge == 0:
                        os.mkdir("./user_json/" + str(message.author.id))
                    
                    # シーズンを分けるファイル
                    files = glob.glob('./user_json/' + str(message.author.id) + '/*')
                    judge = 0

                    for i in range(0, len(files)):
                        #print(os.path.split(files[i])[1])
                        if(os.path.split(files[i])[1] == str(now_season)):
                            judge = 1
                            break
                    
                    # なければ作成
                    if judge == 0:
                        os.mkdir('./user_json/' + str(message.author.id) + '/' + str(now_season))
                    
                    # ndjsonファイルに書き込むものの定義
                    content = {
                        "time" : str(datetime.datetime.now()),
                        "rank": rank
                    }

                    # ndjsonファイルに書き込む
                    with open('./user_json/' + str(message.author.id)  + '/' + str(now_season) + '/' + name.abbr + ".ndjson", 'a') as f:
                        writer = ndjson.writer(f)
                        writer.writerow(content)

                    # 平均順位の算出
                    # ファイルを読み込む
                    with open('./user_json/' + str(message.author.id)  + '/' + str(now_season) + '/' + name.abbr + ".ndjson") as f:
                        read_data = ndjson.load(f)
                    
                    sum = 0
                    avg = 0
                    for i in range(0, len(read_data)):
                        sum += read_data[i]["rank"]
                    
                    avg = sum / len(read_data)

                    # メッセージ表示
                    if language == "ja":
                        embed=discord.Embed(title="\n" +"Season" + str(now_season) +" 順位記録\n\n" + name.full_name_ja,color=0x00ff40)
                    elif language == "en":
                        embed=discord.Embed(title="Ranking register!\n" +"Season" + str(now_season) +" Rank Record\n\n" + name.full_name,color=0x00ff40)
                    embed.set_author(name=message.author.name, icon_url=message.author.avatar)
                    if language == "ja":
                        embed.add_field(name="現在の平均順位", value=str(round(avg, 1)) + "位", inline=True)
                        if len(read_data) >= 2: # 過去の記録が存在するときはこっち
                            embed.add_field(name="前回の順位", value=str(read_data[len(read_data)-2]["rank"]) + "位", inline=True)
                        else: # 記録がないときはこっち
                            embed.add_field(name="前回の順位", value="なし", inline=True)
                    elif language == "en":
                        embed.add_field(name="Now average Rank", value=str(round(avg, 1)), inline=True)
                        if len(read_data) >= 2: # 過去の記録が存在するときはこっち
                            embed.add_field(name="Previous rank", value=str(read_data[len(read_data)-2]["rank"]), inline=True)
                        else: # 記録がないときはこっち
                            embed.add_field(name="Previous rank", value="No data.", inline=True)
                    embed.set_image(url="https://ay2416.github.io/Rank-Collector/stage_picture/" + str(name.id) + ".jpg")
                    embed.set_footer(text="Picture : ©Mario Kart Blog")
                    await message.channel.send(embed=embed)

                    # guild_jsonのtimeを更新
                    # 同じチャンネルのデータがある部分を探す
                    # guild_jsonの読込
                    with open('./guild_json/' + str(message.guild.id) + '.ndjson') as f:
                        read_data = ndjson.load(f)
                    
                    data_location1 = 0
                    for i in range(0, len(read_data)):
                        if read_data[i]["channel"] == message.channel.id:
                            data_location1 = i
                            #print(read_data)
                            break
                    
                    # ndjsonにtimeを書き込み
                    read_data[data_location1]["time"] = str(datetime.datetime.now())

                    os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                    for i in range(0, len(read_data)):
                        with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                            writer = ndjson.writer(f)
                            writer.writerow(read_data[i])
                    
                    # add_jsonのcountを書き換え
                    with open('./add_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson') as f:
                        read_data = ndjson.load(f)

                    now_count = read_data[0]["count"]

                    #print(now_count)
                    # もし12回目の入力であれば終了する
                    if now_count == 12:

                        # 処理が終わったのでadd_jsonを削除する
                        # add_jsonのギルドフォルダ内のフォルダの情報を取得
                        files = glob.glob('./add_json/' + str(message.guild.id) + '/*.ndjson')

                        os.remove('./add_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson')

                        # ファイルの個数が1だった場合はギルドフォルダも削除する
                        if len(files) == 1:
                            os.rmdir('./add_json/' + str(message.guild.id))
                        
                        # guild_jsonの読込
                        with open('./guild_json/' + file) as f:
                            read_data = ndjson.load(f)
                        
                        data_location1 = 0
                        for i in range(0, len(read_data)):
                            if read_data[i]["channel"] == message.channel.id:
                                data_location1 = i
                                #print(read_data)
                                break

                        read_data[data_location1]["mode"] = "null"
                        
                        os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                        for i in range(0, len(read_data)):
                                with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                                    writer = ndjson.writer(f)
                                    writer.writerow(read_data[i])

                        # メッセージ表示
                        if language == "ja":
                            embed=discord.Embed(title="12レース分入力されたためaddモードを終了します！", description="まだ追加したい場合は、\n/addコマンドでstartしてください。", color=0x00ff40)
                        elif language == "en":
                            embed=discord.Embed(title="Exit [add mode] because 12 races have been entered!", description="If you still want to add more, \nstart with the /add command.", color=0x00ff40)                            
                        await message.channel.send(embed=embed)

                        return
                    
                    content = {
                        "count" : now_count + 1,
                    }

                    os.remove('./add_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson')

                    with open('./add_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson', 'a') as f:
                        writer = ndjson.writer(f)
                        writer.writerow(content)

            # モードがdeleteの場合
            elif mode == "delete":

                # 現在のユーザーがどこの手順を行っているかを保存しているものを呼び出し
                with open('./delete_json/' + str(message.guild.id) + '/' + str(message.author.id) + ".ndjson") as f:
                    read_data = ndjson.load(f)
                
                count = read_data[0]["count"]
                
                #print(count)

                # 手順3：続けるか続けないのか
                if count == 2:
                    name = Track.from_nick(read_data[0]["course"])
                    msg = message.content.split()
                    select = str(msg[0])

                    # 続ける場合
                    if select == "Yes" or select == "yes":
                        # 既に登録があるかの確認
                        # ギルドフォルダー
                        files = glob.glob('./delete_json/*')
                        judge = 0

                        for i in range(0, len(files)):
                            #print(os.path.split(files[i])[1])
                            if(os.path.split(files[i])[1] == str(message.guild.id)):
                                judge = 1
                                break
                        
                        # なければ作成
                        if judge == 0:
                            os.mkdir("./delete_json/" + str(message.guild.id))

                        # delete_jsonの作成
                        content = {
                            "count" : 1,
                            "course" : name.abbr
                        }

                        os.remove('./delete_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson')

                        with open('./delete_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson', 'a') as f:
                            writer = ndjson.writer(f)
                            writer.writerow(content)

                        # コースデータの読込
                        with open('./user_json/' + str(message.author.id)  + '/' + str(now_season) + '/' + name.abbr + ".ndjson") as f:
                            read_data = ndjson.load(f)

                        # 表示させる
                        if language == "ja":
                            embed=discord.Embed(title="Season" + str(now_season) +  " 過去記録\n\n" + name.full_name_ja, description="**削除したい記録を番号で選択してください。**\n※直近分表示（最大25個まで）", color=0x00008b)
                        elif language == "en":
                            embed=discord.Embed(title="Season" + str(now_season) +  " Old Record\n\n" + name.full_name, description="**Select the record you wish to delete by number.**\n*Display of most recent minutes (up to 25)", color=0x00008b)        
                        embed.set_author(name=message.author.name, icon_url=message.author.avatar)
                        num = 1
                        for i in range(len(read_data)-1,-1,-1):
                            if language == "ja":
                                date = dateutil.parser.parse(read_data[i]["time"]).astimezone(JST)
                                embed.add_field(name="", value=str(num) + ". " + date.strftime("%Y/%m/%d %H:%M:%S") + " JST : " + str(read_data[i]["rank"]) + "位", inline=False)
                            elif language == "en":
                                date = dateutil.parser.parse(read_data[i]["time"]).astimezone(UTC)
                                embed.add_field(name="", value=str(num) + ". " + date.strftime("%Y/%m/%d %H:%M:%S") + " UTC : " + str(read_data[i]["rank"]), inline=False)
                            num = num + 1
                        embed.set_image(url="https://ay2416.github.io/Rank-Collector/stage_picture/" + str(name.id) + ".jpg")
                        embed.set_footer(text="Picture : ©Mario Kart Blog")
                        
                        await message.channel.send(embed=embed)
                    
                    # 続けない場合
                    elif select == "No" or select == "no":
                        # 既に登録があるかの確認
                        # ギルドフォルダー
                        files = glob.glob('./delete_json/*')
                        judge = 0

                        for i in range(0, len(files)):
                            #print(os.path.split(files[i])[1])
                            if(os.path.split(files[i])[1] == str(message.guild.id)):
                                judge = 1
                                break
                        
                        # なければ作成
                        if judge == 0:
                            os.mkdir("./delete_json/" + str(message.guild.id))

                        #delete_jsonの作成
                        content = {
                            "count" : 0
                        }

                        os.remove('./delete_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson')

                        with open('./delete_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson', 'a') as f:
                            writer = ndjson.writer(f)
                            writer.writerow(content)

                        '''
                        #コースデータの読込 多分この処理いらないはず　消しても良し
                        with open('./user_json/' + str(message.author.id)  + '/' + str(now_season) + '/' + name.abbr + ".ndjson") as f:
                            read_data = ndjson.load(f)
                        '''
                        if language == "ja":
                            embed=discord.Embed(title="終了しました！", description="別のコースの記録の削除を行いたい場合はコース名を入力してください。\n削除を終了したい場合は、/deleteコマンドでstopを実行してください。", color=0x00ff40)
                        elif language == "en":
                            embed=discord.Embed(title="Completed!", description="If you want to delete a record from another course, enter the name of the course.\nIf you wish to terminate the deletion, execute stop with the /delete command.", color=0x00ff40)
                        await message.channel.send(embed=embed)
                    
                    # 該当する文字が入力されなかった場合
                    else:
                        if language == "ja":
                            embed=discord.Embed(title="入力された文字が間違っています！", color=0xff0000)
                            embed.add_field(name="Yes/No（yes/no）で解答してください。", value="", inline=False)
                        elif language == "en":
                            embed=discord.Embed(title="The entered characters are incorrect!", color=0xff0000)
                            embed.add_field(name="Please answer with Yes/No (yes/no).", value="", inline=False)
                        await message.channel.send(embed=embed)
                    
                        # guild_jsonのtimeを更新
                        # 同じチャンネルのデータがある部分を探す
                        # guild_jsonの読込
                        with open('./guild_json/' + str(message.guild.id) + '.ndjson') as f:
                            read_data = ndjson.load(f)
                        
                        data_location1 = 0
                        for i in range(0, len(read_data)):
                            if read_data[i]["channel"] == message.channel.id:
                                data_location1 = i
                                #print(read_data)
                                break
                        
                        # ndjsonにtimeを書き込み
                        read_data[data_location1]["time"] = str(datetime.datetime.now())

                        os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                        for i in range(0, len(read_data)):
                            with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                                writer = ndjson.writer(f)
                                writer.writerow(read_data[i])
                
                # 手順2：削除する〇番目の指定
                elif count == 1:
                    name = Track.from_nick(read_data[0]["course"])
                    msg = message.content.split()
                    
                    #print(msg[0])
                    
                    num = int(msg[0])
                    
                    # コースデータの読込
                    with open('./user_json/' + str(message.author.id)  + '/' + str(now_season) + '/' + name.abbr + ".ndjson") as f:
                        read_data = ndjson.load(f)

                    # 入力された数が範囲内であるかの判定
                    if num < 1 or num > 25:
                        if language == "ja":
                            embed=discord.Embed(title="入力された数が有効範囲外です！", color=0xff0000)
                            embed.add_field(name="有効な数の範囲は1~25です。", value=" ", inline=False)
                        elif language == "en":
                            embed=discord.Embed(title="The entered rank is out of the valid range!", color=0xff0000)
                            embed.add_field(name="The valid ranking range is 1~12.", value="", inline=False)
                        await message.channel.send(embed=embed)

                        # guild_jsonのtimeを更新
                        # 同じチャンネルのデータがある部分を探す
                        # guild_jsonの読込
                        with open('./guild_json/' + str(message.guild.id) + '.ndjson') as f:
                            read_data = ndjson.load(f)
                        
                        data_location1 = 0
                        for i in range(0, len(read_data)):
                            if read_data[i]["channel"] == message.channel.id:
                                data_location1 = i
                                #print(read_data)
                                break
                        
                        # ndjsonにtimeを書き込み
                        read_data[data_location1]["time"] = str(datetime.datetime.now())

                        os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                        for i in range(0, len(read_data)):
                            with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                                writer = ndjson.writer(f)
                                writer.writerow(read_data[i])
                        
                        return
                    
                    # コースデータの更新
                    if len(read_data) == 1:
                        os.remove('./user_json/' + str(message.author.id)  + '/' + str(now_season) + '/' + name.abbr + ".ndjson")
                        if language == "ja":
                            embed=discord.Embed(title="削除が完了しました！", description="このコースのデータがすべて削除されたため、このコースの削除の処理を終了します。\n別のコースの削除を行いたい場合はコース名を入力してください。\n削除を終了したい場合は/deleteコマンドでstopをしてください。",color=0x00ff40)
                        elif language == "en":
                            embed=discord.Embed(title="Deletion completed!", description="The process of deleting this course is terminated because all data for this course has been deleted.\nIf you want to delete another course, enter the name of the course.\nIf you want to finish deleting the course, use the /delete command to stop.",color=0x00ff40)
                        await message.channel.send(embed=embed)

                        # 既に登録があるかの確認
                        # ギルドフォルダー
                        files = glob.glob('./delete_json/*')
                        judge = 0

                        for i in range(0, len(files)):
                            #print(os.path.split(files[i])[1])
                            if(os.path.split(files[i])[1] == str(message.guild.id)):
                                judge = 1
                                break
                        
                        # なければ作成
                        if judge == 0:
                            os.mkdir("./delete_json/" + str(message.guild.id))

                        # delete_jsonの作成
                        content = {
                            "count" : 0
                        }

                        os.remove('./delete_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson')

                        with open('./delete_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson', 'a') as f:
                            writer = ndjson.writer(f)
                            writer.writerow(content)

                    else:
                        data_location = 0
                        for i in range(len(read_data)-1,-1,-1):
                            if i == len(read_data)- num:
                                data_location = i
                                #print(read_data)
                                break

                        #if data_write == 1:
                        os.remove('./user_json/' + str(message.author.id)  + '/' + str(now_season) + '/' + name.abbr + ".ndjson")

                        for i in range(0,len(read_data)):
                            if i != data_location:
                                with open('./user_json/' + str(message.author.id)  + '/' + str(now_season) + '/' + name.abbr + ".ndjson", 'a') as f:
                                    writer = ndjson.writer(f)
                                    writer.writerow(read_data[i])
                        if language == "ja":
                            embed=discord.Embed(title="削除が完了しました！", description="まだ同じコースの記録の削除を続けますか？（Yes/No）",color=0x00ff40)
                        elif language == "en":
                            embed=discord.Embed(title="Deletion completed!", description="Do you still want to continue deleting records for the same course? (Yes/No)",color=0x00ff40)                            
                        await message.channel.send(embed=embed)
                    
                        # 既に登録があるかの確認
                        # ギルドフォルダー
                        files = glob.glob('./delete_json/*')
                        judge = 0

                        for i in range(0, len(files)):
                            #print(os.path.split(files[i])[1])
                            if(os.path.split(files[i])[1] == str(message.guild.id)):
                                judge = 1
                                break
                        
                        # なければ作成
                        if judge == 0:
                            os.mkdir("./delete_json/" + str(message.guild.id))

                        # delete_jsonの作成
                        content = {
                            "count" : 2,
                            "course" : name.abbr
                        }

                        os.remove('./delete_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson')

                        with open('./delete_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson', 'a') as f:
                            writer = ndjson.writer(f)
                            writer.writerow(content)

                        # guild_jsonのtimeを更新
                        # 同じチャンネルのデータがある部分を探す
                        # guild_jsonの読込
                        with open('./guild_json/' + str(message.guild.id) + '.ndjson') as f:
                            read_data = ndjson.load(f)
                        
                        data_location1 = 0
                        for i in range(0, len(read_data)):
                            if read_data[i]["channel"] == message.channel.id:
                                data_location1 = i
                                #print(read_data)
                                break
                        
                        # ndjsonにtimeを書き込み
                        read_data[data_location1]["time"] = str(datetime.datetime.now())

                        os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                        for i in range(0, len(read_data)):
                            with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                                writer = ndjson.writer(f)
                                writer.writerow(read_data[i])

                # 手順1：削除するコースの指定      
                else:
                    msg = message.content.split()
                    
                    #print(msg[0])
                    
                    course = str(msg[0])

                    # 入力されたコース名がデータにあるかの判定
                    try:
                        name = Track.from_nick(course)
                        print(name.abbr)
                    except Exception as e:
                        if language == "ja":
                            embed=discord.Embed(title="そのコースデータは存在しません！", color=0xff0000)
                            embed.add_field(name="入力されたコース名を確認してください。", value=" ", inline=False)
                        elif language == "en":
                            embed=discord.Embed(title="That course data does not exist!", color=0xff0000)
                            embed.add_field(name="Please confirm the course name entered.", value=" ", inline=False)
                        await message.channel.send(embed=embed)

                        # guild_jsonのtimeを更新
                        # 同じチャンネルのデータがある部分を探す
                        # guild_jsonの読込
                        with open('./guild_json/' + str(message.guild.id) + '.ndjson') as f:
                            read_data = ndjson.load(f)
                        
                        data_location1 = 0
                        for i in range(0, len(read_data)):
                            if read_data[i]["channel"] == message.channel.id:
                                data_location1 = i
                                #print(read_data)
                                break
                        
                        # ndjsonにtimeを書き込み
                        read_data[data_location1]["time"] = str(datetime.datetime.now())

                        os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                        for i in range(0, len(read_data)):
                            with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                                writer = ndjson.writer(f)
                                writer.writerow(read_data[i])
                        
                        return
                    
                    # コースデータが存在しているかの確認
                    # コースを保存するユーザーデータフォルダがあるか
                    files = glob.glob('./user_json/*')
                    judge = 0

                    for i in range(0, len(files)):
                        #print(os.path.split(files[i])[1])
                        if(os.path.split(files[i])[1] == str(message.author.id)):
                            judge = 1
                            break
                    
                    # なければエラー
                    if judge == 0:
                        if language == "ja":
                            embed=discord.Embed(title="あなたは順位の記録を登録していません！", color=0xff0000)
                            embed.add_field(name="順位の記録を登録してください。", value=" ", inline=False)
                        elif language == "en":
                            embed=discord.Embed(title="You have not registered your rank record!", color=0xff0000)
                            embed.add_field(name="Register your ranking record.", value=" ", inline=False)
                        await message.channel.send(embed=embed)

                        # guild_jsonのtimeを更新
                        # 同じチャンネルのデータがある部分を探す
                        # guild_jsonの読込
                        with open('./guild_json/' + str(message.guild.id) + '.ndjson') as f:
                            read_data = ndjson.load(f)
                        
                        data_location1 = 0
                        for i in range(0, len(read_data)):
                            if read_data[i]["channel"] == message.channel.id:
                                data_location1 = i
                                #print(read_data)
                                break
                        
                        # ndjsonにtimeを書き込み
                        read_data[data_location1]["time"] = str(datetime.datetime.now())

                        os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                        for i in range(0, len(read_data)):
                            with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                                writer = ndjson.writer(f)
                                writer.writerow(read_data[i])
                        
                        return
                    
                    # seasonフォルダが作成されているか
                    files = glob.glob('./user_json/' + str(message.author.id) + '/*')
                    judge = 0

                    for i in range(0, len(files)):
                        #print(os.path.split(files[i])[1])
                        if(os.path.split(files[i])[1] == str(now_season)):
                            judge = 1
                            break
                    
                    # なければエラー
                    if judge == 0:
                        if language == "ja":
                            embed=discord.Embed(title="あなたはシーズンが変わってから順位の記録をしていません！", color=0xff0000)
                            embed.add_field(name="順位の記録を登録してください。", value=" ", inline=False)
                        elif language == "en":
                            embed=discord.Embed(title="You haven't recorded your standings since the season changed!", color=0xff0000)
                            embed.add_field(name="Register your ranking record.", value=" ", inline=False)
                        await message.channel.send(embed=embed)

                        # guild_jsonのtimeを更新
                        # 同じチャンネルのデータがある部分を探す
                        # guild_jsonの読込
                        with open('./guild_json/' + str(message.guild.id) + '.ndjson') as f:
                            read_data = ndjson.load(f)
                        
                        data_location1 = 0
                        for i in range(0, len(read_data)):
                            if read_data[i]["channel"] == message.channel.id:
                                data_location1 = i
                                #print(read_data)
                                break
                        
                        # ndjsonにtimeを書き込み
                        read_data[data_location1]["time"] = str(datetime.datetime.now())

                        os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                        for i in range(0, len(read_data)):
                            with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                                writer = ndjson.writer(f)
                                writer.writerow(read_data[i])
                        
                        return
                    
                    # コース名.ndjsonがあるか
                    files = glob.glob('./user_json/' + str(message.author.id)  + '/' + str(now_season) + '/*.ndjson')
                    judge = 0

                    for i in range(0, len(files)):
                        #print(os.path.split(files[i])[1])
                        if(os.path.split(files[i])[1] == name.abbr + '.ndjson'):
                            judge = 1
                            break
                    
                    # なければエラー
                    if judge == 0:
                        if language == "ja":
                            embed=discord.Embed(title="そのコースの順位の記録は存在しません！", color=0xff0000)
                            embed.add_field(name="そのコースの順位の記録を登録してください。", value=" ", inline=False)
                        elif language == "en":
                            embed=discord.Embed(title="No record of the rankings for that course exists!", color=0xff0000)
                            embed.add_field(name="Register a record of your rank for that course.", value=" ", inline=False)
                        await message.channel.send(embed=embed)

                        # guild_jsonのtimeを更新
                        # 同じチャンネルのデータがある部分を探す
                        # guild_jsonの読込
                        with open('./guild_json/' + str(message.guild.id) + '.ndjson') as f:
                            read_data = ndjson.load(f)
                        
                        data_location1 = 0
                        for i in range(0, len(read_data)):
                            if read_data[i]["channel"] == message.channel.id:
                                data_location1 = i
                                #print(read_data)
                                break
                        
                        # ndjsonにtimeを書き込み
                        read_data[data_location1]["time"] = str(datetime.datetime.now())

                        os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                        for i in range(0, len(read_data)):
                            with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                                writer = ndjson.writer(f)
                                writer.writerow(read_data[i])
                        
                        return

                    #コースデータの読込
                    with open('./user_json/' + str(message.author.id)  + '/' + str(now_season) + '/' + name.abbr + ".ndjson") as f:
                        read_data = ndjson.load(f)

                    # 表示させる
                    if language == "ja":
                        embed=discord.Embed(title="Season" + str(now_season) +  " 過去記録\n\n" + name.full_name_ja, description="**削除したい記録を番号で選択してください。**\n※直近分表示（最大25個まで）", color=0x00008b)
                    elif language == "en":
                        embed=discord.Embed(title="Season" + str(now_season) +  " Old Record\n\n" + name.full_name, description="**Select the record you wish to delete by number.**\n*Display of most recent minutes (up to 25)", color=0x00008b)        
                    embed.set_author(name=message.author.name, icon_url=message.author.avatar)
                    num = 1
                    for i in range(len(read_data)-1,-1,-1):
                        if language == "ja":
                            date = dateutil.parser.parse(read_data[i]["time"]).astimezone(JST)
                            embed.add_field(name="", value=str(num) + ". " + date.strftime("%Y/%m/%d %H:%M:%S") + " JST : " + str(read_data[i]["rank"]) + "位", inline=False)
                        elif language == "en":
                            date = dateutil.parser.parse(read_data[i]["time"]).astimezone(UTC)
                            embed.add_field(name="", value=str(num) + ". " + date.strftime("%Y/%m/%d %H:%M:%S") + " UTC : " + str(read_data[i]["rank"]), inline=False)
                        num = num + 1
                    embed.set_image(url="https://ay2416.github.io/Rank-Collector/stage_picture/" + str(name.id) + ".jpg")
                    embed.set_footer(text="Picture : ©Mario Kart Blog")
                    
                    await message.channel.send(embed=embed)


                    '''
                    50以上の数字が50のままにあるバグあり
                    embed=discord.Embed(title="Season" + str(now_season) +  " 過去記録\n\n" + name.full_name_ja, description="**削除したい記録を番号で選択してください。**", color=0x00008b)
                    embed.set_author(name=message.author.name, icon_url=message.author.avatar)
                    cut = 25
                    for i in range(0, len(read_data)):
                        
                        date = dateutil.parser.parse(read_data[i]["time"]).astimezone(JST)
                        embed.add_field(name=" ", value=str(i) + ". " + date.strftime("%Y/%m/%d %H:%M:%S") + " JST : " + str(read_data[i]["rank"]) + "位", inline=False)
                        if i == len(read_data) - 1:
                            #表示させる
                            embed.set_image(url="https://ay2416.github.io/Rank-Collector/stage_picture/" + str(name.id) + ".jpg")
                            embed.set_footer(text="Picture : ©Mario Kart Blog")
                            await message.channel.send(embed=embed)
                            return
                        
                        if i + 1 == cut:
                            cut = cut + 25
                            #表示させる
                            await message.channel.send(embed=embed)
                            await asyncio.sleep(2)
                            embed=discord.Embed(title="", color=0x00008b)
                    '''
                
                    # 既に登録があるかの確認
                    # ギルドフォルダー
                    files = glob.glob('./delete_json/*')
                    judge = 0

                    for i in range(0, len(files)):
                        #print(os.path.split(files[i])[1])
                        if(os.path.split(files[i])[1] == str(message.guild.id)):
                            judge = 1
                            break
                    
                    # なければ作成
                    if judge == 0:
                        os.mkdir("./delete_json/" + str(message.guild.id))

                    #delete_jsonの作成
                    content = {
                        "count" : 1,
                        "course" : name.abbr
                    }

                    os.remove('./delete_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson')

                    with open('./delete_json/' + str(message.guild.id) + '/' + str(message.author.id) + '.ndjson', 'a') as f:
                        writer = ndjson.writer(f)
                        writer.writerow(content)

                    # guild_jsonのtimeを更新
                    # 同じチャンネルのデータがある部分を探す
                    # guild_jsonの読込
                    with open('./guild_json/' + str(message.guild.id) + '.ndjson') as f:
                        read_data = ndjson.load(f)
                    
                    data_location1 = 0
                    for i in range(0, len(read_data)):
                        if read_data[i]["channel"] == message.channel.id:
                            data_location1 = i
                            #print(read_data)
                            break
                    
                    # ndjsonにtimeを書き込み
                    read_data[data_location1]["time"] = str(datetime.datetime.now())

                    os.remove('./guild_json/' + str(message.guild.id) + '.ndjson')

                    for i in range(0, len(read_data)):
                        with open('./guild_json/' + str(message.guild.id) + '.ndjson', 'a') as f:
                            writer = ndjson.writer(f)
                            writer.writerow(read_data[i])

    except Exception as e:
        pass

# エラーが出ない限り、10秒ごとに実行し続ける
@tasks.loop(seconds=60)
async def time_check():
    # guild_jsonのファイル一覧を取得
    files = glob.glob('guild_json/*.ndjson')

    # add,deleteモードの時に指定の時間数が経っていた場合に処理を終了する処理
    for i in range(0, len(files)):
        with open('./guild_json/' + os.path.split(files[i])[1]) as f:
            read_data = ndjson.load(f)

        # 言語の確認
        file = os.path.split(files[i])[1]

        with open('./language_json/' + file) as f:
            read_data2 = ndjson.load(f)

        language = read_data2[0]["language_mode"]

        for j in range(0, len(read_data)):
            if read_data[j]["mode"] == "add" or read_data[j]["mode"] == "delete":
                time = datetime.datetime.strptime(read_data[j]["time"], time_format)

                if (datetime.datetime.now() - time).total_seconds() >= stop_time:
                    read_data[j]["mode"] = "null"
                    
                    os.remove('./guild_json/' + os.path.split(files[i])[1])

                    for k in range(0, len(read_data)):
                            with open('./guild_json/' + os.path.split(files[i])[1], 'a') as f:
                                writer = ndjson.writer(f)
                                writer.writerow(read_data[i])

                    if language == "ja":
                        embed=discord.Embed(title="10分間操作がなかったため処理を終了します！", description="また使い場合はコマンドを実行してください。", color=0x00008b)
                    elif language == "en":
                        embed=discord.Embed(title="The process is terminated because there has been no operation for 10 minutes!", description="If you want to use it again, please execute the command.", color=0x00008b)
                    
                    channel_sent = client.get_channel(read_data[j]["channel"])
                    await channel_sent.send(embed=embed)

#Bot commands
# /add
@tree.command(name="add",description="順位の記録の追加を開始・停止します。 / Rank data add mode change [start] and [stop].")
@discord.app_commands.guild_only()
@discord.app_commands.choices(mode=[discord.app_commands.Choice(name="start",value="start"),discord.app_commands.Choice(name="stop",value="stop")])
async def add_command(interaction: discord.Interaction,mode:str):
    
    await interaction.response.defer(ephemeral=False)

    # 言語の確認
    file = str(interaction.guild.id) + ".ndjson"

    with open('./language_json/' + file) as f:
        read_data = ndjson.load(f)

    language = read_data[0]["language_mode"]

    # コマンドプログラム本体

    # モードにstartが入力された場合
    if mode == "start":

        # modeを保存する処理
        # guild_jsonが存在しているかの確認
        files = glob.glob('./guild_json/*.ndjson')
        judge = 0

        for i in range(0, len(files)):
            #print(os.path.split(files[i])[1])
            if os.path.split(files[i])[1] == str(interaction.guild.id) + ".ndjson":
                judge = 1
                break 
        
        # あった場合の処理
        if judge == 1:
            # 同じチャンネルのデータがあるかの確認
            # guild_jsonの読込
            with open('./guild_json/' + str(interaction.guild.id) + '.ndjson') as f:
                read_data = ndjson.load(f)
            
            judge = 0
            data_location = 0
            for i in range(0, len(read_data)):
                if read_data[i]["channel"] == interaction.channel.id:
                    judge = 1
                    data_location = i
                    #print(read_data)
                    break
            
            if judge == 1:
                # modeの状態を代入
                data = read_data[data_location]["mode"]

                # 既にそのモードに入っていたり、別のモードになっていないかを確認（誤動作防止）
                if data == "add":
                    if language == "ja":
                        embed=discord.Embed(title="既にaddモードに入っています！", description="終了させたい場合は/add stopを実行してください。", color=0xff0000)
                    elif language == "en":
                        embed=discord.Embed(title="We are already in add mode!", description="If you want to terminate the program, execute /add stop.", color=0xff0000)
                    await interaction.followup.send(embed=embed)
                    return
                elif data == "delete":
                    if language == "ja":
                        embed=discord.Embed(title="deleteモードに入っています！", description="こちらのモードを使用したい場合は/delete stopで終了してからもう一度実行してください。", color=0xff0000)
                    elif language == "en":
                        embed=discord.Embed(title="You are in delete mode!", description="If you want to use this mode, exit with /delete stop and run it again.", color=0xff0000)
                    await interaction.followup.send(embed=embed)
                    return

                read_data[data_location]["mode"] = "add"
                read_data[data_location]["time"] = str(datetime.datetime.now())

                os.remove('./guild_json/' + str(interaction.guild.id) + '.ndjson')

                for i in range(0, len(read_data)):
                    with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                        writer = ndjson.writer(f)
                        writer.writerow(read_data[i])
            else:
                content = {
                    "mode" : "add",
                    "channel" : interaction.channel.id,
                    "time" : str(datetime.datetime.now())
                }

                with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                    writer = ndjson.writer(f)
                    writer.writerow(content)
        
        # なかった場合の処理
        else:
            content = {
                "mode" : "add",
                "channel" : interaction.channel.id,
                "time" : str(datetime.datetime.now())
            }

            with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                writer = ndjson.writer(f)
                writer.writerow(content)
        
        # add_jsonフォルダの
        # ギルドフォルダーの確認
        files = glob.glob('./add_json/*')
        judge = 0

        for i in range(0, len(files)):
            #print(os.path.split(files[i])[1])
            if(os.path.split(files[i])[1] == str(interaction.guild.id)):
                judge = 1
                break
        
        # なければ作成
        if judge == 0:
            os.mkdir("./add_json/" + str(interaction.guild.id))

        #delete_jsonの作成
        content = {
            "count" : 1
        }

        # ファイルが存在しなくてエラーが発生した時はファイルの削除を無視して進む
        try:
            os.remove('./add_json/' + str(interaction.guild.id) + '/' + str(interaction.user.id) + '.ndjson')
        except Exception as e:
            pass

        with open('./add_json/' + str(interaction.guild.id) + '/' + str(interaction.user.id) + '.ndjson', 'a') as f:
            writer = ndjson.writer(f)
            writer.writerow(content)
         
        if language == "ja":
            embed=discord.Embed(title="順位の登録を開始します！", description="「[コース名] [順位]」の形式でテキストを入力してください。\n終了したい場合は、もう一度/addコマンドでモードをstopにしてください。",color=0x00ff40)
        elif language == "en":
            embed=discord.Embed(title="Starts the registration of rankings!", description="Enter text in the format 【[course name] [rank]】.\nIf you want to exit, use the /add command again to set the mode to stop.",color=0x00ff40)            
        await interaction.followup.send(embed=embed)        

    # モードにstopが入力された場合
    elif mode == "stop":

        # modeを保存する処理
        # guild_jsonが存在しているかの確認
        files = glob.glob('./guild_json/*.ndjson')
        judge = 0

        for i in range(0, len(files)):
            #print(os.path.split(files[i])[1])
            if os.path.split(files[i])[1] == str(interaction.guild.id) + ".ndjson":
                judge = 1
                break
        
        # あった場合の処理
        if judge == 1:
            # 同じチャンネルのデータがあるかの確認
            # guild_jsonの読込
            with open('./guild_json/' + str(interaction.guild.id) + '.ndjson') as f:
                read_data = ndjson.load(f)
            
            judge = 0
            data_location = 0
            for i in range(0, len(read_data)):
                if read_data[i]["channel"] == interaction.channel.id:
                    judge = 1
                    data_location = i
                    #print(read_data)
                    break
            
            if judge == 1:
                # modeの状態を代入
                data = read_data[data_location]["mode"]

                # 既にそのモードに入っていたり、別のモードになっていないかを確認（誤動作防止）
                if data == "null":
                    if language == "ja":
                        embed=discord.Embed(title="既にaddモードを終了しています！", description="開始したい場合は/add startを実行してください。", color=0xff0000)
                    elif language == "en":
                        embed=discord.Embed(title="I have already exited add mode!", description="If you want to start, execute /add start.", color=0xff0000)
                    await interaction.followup.send(embed=embed)
                    return
                elif data == "delete":
                    if language == "ja":
                        embed=discord.Embed(title="deleteモードに入っています！", description="こちらのモードを使用したい場合は/delete stopで終了してからもう一度実行してください。", color=0xff0000)
                    elif language == "en":
                        embed=discord.Embed(title="You are in delete mode!", description="If you want to use this mode, exit with /delete stop and run it again.", color=0xff0000)
                    await interaction.followup.send(embed=embed)
                    return

                read_data[data_location]["mode"] = "null"
                read_data[data_location]["time"] = str(datetime.datetime.now())

                os.remove('./guild_json/' + str(interaction.guild.id) + '.ndjson')

                for i in range(0, len(read_data)):
                    with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                        writer = ndjson.writer(f)
                        writer.writerow(read_data[i])
            else:
                content = {
                    "mode" : "null",
                    "channel" : interaction.channel.id
                }

                with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                    writer = ndjson.writer(f)
                    writer.writerow(content)
        
        # なかった場合の処理
        else:
            content = {
                "mode" : "null",
                "channel" : interaction.channel.id
            }

            with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                writer = ndjson.writer(f)
                writer.writerow(content)
        
        # 処理が終わったのでadd_jsonを削除する
        # add_jsonのギルドフォルダ内のフォルダの情報を取得
        files = glob.glob('./add_json/' + str(interaction.guild.id) + '/*.ndjson')

        os.remove('./add_json/' + str(interaction.guild.id) + '/' + str(interaction.user.id) + '.ndjson')

        # ファイルの個数が1だった場合はギルドフォルダも削除する
        if len(files) == 1:
            os.rmdir('./add_json/' + str(interaction.guild.id))

        if language == "ja":    
            embed=discord.Embed(title="順位の登録を終了しました！", description="また使用したい場合は/addコマンドでモードをstartに切り替えてください。",color=0x00ff40)
        elif language == "en":
            embed=discord.Embed(title="Registration for the ranking has been closed!", description="If you want to use it again, use the /add command to switch the mode to start.",color=0x00ff40)
        
        await interaction.followup.send(embed=embed)
        
    # それ以外の者が入力された時
    else:
        if language == "ja":
            embed=discord.Embed(title="入力されたモードがありません！", description="入力されたモードを確認してください。", color=0xff0000)
        elif language == "en":
            embed=discord.Embed(title="No mode entered!", description="Check the mode entered.", color=0xff0000)
        
        await interaction.followup.send(embed=embed)
  
# /delete
@tree.command(name="delete",description="順位の記録の削除を開始・停止します。 / Rank data delete mode change [start] and [stop]..")
@discord.app_commands.guild_only()
@discord.app_commands.choices(mode=[discord.app_commands.Choice(name="start",value="start"),discord.app_commands.Choice(name="stop",value="stop")])
async def delete_command(interaction: discord.Interaction,mode:str):

    await interaction.response.defer(ephemeral=False)

    # 言語の確認
    file = str(interaction.guild.id) + ".ndjson"

    with open('./language_json/' + file) as f:
        read_data = ndjson.load(f)

    language = read_data[0]["language_mode"]

    # コマンドプログラム本体

    # モードにstartが入力された場合
    if mode == "start":

        # modeを保存する処理
        # guild_jsonが存在しているかの確認
        files = glob.glob('./guild_json/*.ndjson')
        judge = 0

        for i in range(0, len(files)):
            #print(os.path.split(files[i])[1])
            if os.path.split(files[i])[1] == str(interaction.guild.id) + ".ndjson":
                judge = 1
                break 
        
        # あった場合の処理
        if judge == 1:
            # 同じチャンネルのデータがあるかの確認
            # guild_jsonの読込
            with open('./guild_json/' + str(interaction.guild.id) + '.ndjson') as f:
                read_data = ndjson.load(f)
            
            judge = 0
            data_location = 0
            for i in range(0, len(read_data)):
                if read_data[i]["channel"] == interaction.channel.id:
                    judge = 1
                    data_location = i
                    #print(read_data)
                    break
            
            if judge == 1:
                # modeの状態を代入
                data = read_data[data_location]["mode"]

                # 既にそのモードに入っていたり、別のモードになっていないかを確認（誤動作防止）
                if data == "delete":
                    if language == "ja":
                        embed=discord.Embed(title="既にdeleteモードに入っています！", description="終了させたい場合は/delete stopを実行してください。", color=0xff0000)
                    elif language == "en":
                        embed=discord.Embed(title="It is already in delete mode!", description="If you want to terminate the program, execute /delete stop.", color=0xff0000)

                    await interaction.followup.send(embed=embed)
                    return
                elif data == "add":
                    if language == "ja":
                        embed=discord.Embed(title="addモードに入っています！", description="こちらのモードを使用したい場合は/add stopで終了してからもう一度実行してください。", color=0xff0000)
                    elif language == "en":
                        embed=discord.Embed(title="It is in add mode!", description="If you want to use this mode, exit with /add stop and run it again.", color=0xff0000)

                    await interaction.followup.send(embed=embed)
                    return

                read_data[data_location]["mode"] = "delete"
                read_data[data_location]["time"] = str(datetime.datetime.now())

                os.remove('./guild_json/' + str(interaction.guild.id) + '.ndjson')

                for i in range(0, len(read_data)):
                    with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                        writer = ndjson.writer(f)
                        writer.writerow(read_data[i])
            else:
                content = {
                    "mode" : "delete",
                    "channel" : interaction.channel.id,
                    "time" : str(datetime.datetime.now())
                }

                with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                    writer = ndjson.writer(f)
                    writer.writerow(content)
        
        # なかった場合の処理
        else:
            content = {
                "mode" : "delete",
                "channel" : interaction.channel.id,
                "time" : str(datetime.datetime.now())
            }

            with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                writer = ndjson.writer(f)
                writer.writerow(content)

        # delete_jsonフォルダの
        # ギルドフォルダーの確認
        files = glob.glob('./delete_json/*')
        judge = 0

        for i in range(0, len(files)):
            #print(os.path.split(files[i])[1])
            if(os.path.split(files[i])[1] == str(interaction.guild.id)):
                judge = 1
                break
        
        # なければ作成
        if judge == 0:
            os.mkdir("./delete_json/" + str(interaction.guild.id))

        #delete_jsonの作成
        content = {
            "count" : 0
        }

        # ファイルが存在しなくてエラーが発生した時はファイルの削除を無視して進む
        try:
            os.remove('./delete_json/' + str(interaction.guild.id) + '/' + str(interaction.user.id) + '.ndjson')
        except Exception as e:
            pass

        with open('./delete_json/' + str(interaction.guild.id) + '/' + str(interaction.user.id) + '.ndjson', 'a') as f:
            writer = ndjson.writer(f)
            writer.writerow(content) 
        
        if language == "ja":
            embed=discord.Embed(title="順位の削除を開始します！", description="コース名をテキストで入力してください。\n終了したい場合は、もう一度/deleteコマンドでモードをstopにしてください。",color=0x00ff40)
        elif language == "en":
            embed=discord.Embed(title="Starts the deletion of rankings!", description="Enter the course name in text.\nIf you want to exit, use the /delete command again to set the mode to stop.",color=0x00ff40)
        
        await interaction.followup.send(embed=embed)        

    # モードにstopが入力された場合
    elif mode == "stop":

        # modeを保存する処理
        # guild_jsonが存在しているかの確認
        files = glob.glob('./guild_json/*.ndjson')
        judge = 0

        for i in range(0, len(files)):
            #print(os.path.split(files[i])[1])
            if os.path.split(files[i])[1] == str(interaction.guild.id) + ".ndjson":
                judge = 1
                break
        
        # あった場合の処理
        if judge == 1:
            # 同じチャンネルのデータがあるかの確認
            # guild_jsonの読込
            with open('./guild_json/' + str(interaction.guild.id) + '.ndjson') as f:
                read_data = ndjson.load(f)
            
            judge = 0
            data_location = 0
            for i in range(0, len(read_data)):
                if read_data[i]["channel"] == interaction.channel.id:
                    judge = 1
                    data_location = i
                    #print(read_data)
                    break
            
            if judge == 1:
                # modeの状態を代入
                data = read_data[data_location]["mode"]

                # 既にそのモードに入っていたり、別のモードになっていないかを確認（誤動作防止）
                if data == "null":
                    if language == "ja":
                        embed=discord.Embed(title="既にdeleteモードを終了しています！", description="開始したい場合は/delete startを実行してください。", color=0xff0000)
                    elif language == "en":
                        embed=discord.Embed(title="You have already exited delete mode!", description="If you want to start, execute /delete start.", color=0xff0000)
                    
                    await interaction.followup.send(embed=embed)
                    return
                elif data == "add":
                    if language == "ja":
                        embed=discord.Embed(title="addモードに入っています！", description="こちらのモードを使用したい場合は/delete stopで終了してからもう一度実行してください。", color=0xff0000)
                    elif language == "en":
                        embed=discord.Embed(title="It is in add mode!", description="If you want to use this mode, exit with /delete stop and run it again.", color=0xff0000)
                    await interaction.followup.send(embed=embed)
                    return

                read_data[data_location]["mode"] = "null"
                read_data[data_location]["time"] = str(datetime.datetime.now())

                os.remove('./guild_json/' + str(interaction.guild.id) + '.ndjson')

                for i in range(0, len(read_data)):
                    with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                        writer = ndjson.writer(f)
                        writer.writerow(read_data[i])
            else:
                content = {
                    "mode" : "null",
                    "channel" : interaction.channel.id
                }

                with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                    writer = ndjson.writer(f)
                    writer.writerow(content)
        
        # なかった場合の処理
        else:
            content = {
                "mode" : "null",
                "channel" : interaction.channel.id
            }

            with open('./guild_json/' + str(interaction.guild.id) + '.ndjson', 'a') as f:
                writer = ndjson.writer(f)
                writer.writerow(content)
        
        # 処理が終わったのでdelete_jsonを削除する
        # delete_jsonのギルドフォルダ内のフォルダの情報を取得
        files = glob.glob('./delete_json/' + str(interaction.guild.id) + '/*.ndjson')

        os.remove('./delete_json/' + str(interaction.guild.id) + '/' + str(interaction.user.id) + '.ndjson')

        # ファイルの個数が1だった場合はギルドフォルダも削除する
        if len(files) == 1:
            os.rmdir('./delete_json/' + str(interaction.guild.id))

        if language == "ja":
            embed=discord.Embed(title="順位の削除を終了しました！", description="また使用したい場合は/deleteコマンドでモードをstartに切り替えてください。",color=0x00ff40)
        elif language == "en":
            embed=discord.Embed(title="The deletion of rankings has been completed!", description="If you want to use it again, use the /delete command to switch the mode to start.",color=0x00ff40)            

        await interaction.followup.send(embed=embed)
    
    # それ以外の者が入力された時
    else:
        if language == "ja":
            embed=discord.Embed(title="入力されたモードがありません！", description="入力されたモードを確認してください。", color=0xff0000)
        elif language == "en":
            embed=discord.Embed(title="No mode entered!", description="Check the mode entered.", color=0xff0000)
        
        await interaction.followup.send(embed=embed)

# /list_avg
@tree.command(name="list_avg",description="今までに入力された順位の平均順位をコースごとに表示します。 / Display all course average rank.")
@discord.app_commands.guild_only()
@discord.app_commands.choices(display_mode=[discord.app_commands.Choice(name="降順 / desc (12→1)",value="desc"),discord.app_commands.Choice(name="昇順 / asc (1→12)",value="asc")])
@discord.app_commands.choices(season=[discord.app_commands.Choice(name="9",value="9")])
async def list_avg_command(interaction: discord.Interaction,display_mode:str,season:str = None):
    # 言語の確認
    file = str(interaction.guild.id) + ".ndjson"

    with open('./language_json/' + file) as f:
        read_data = ndjson.load(f)

    language = read_data[0]["language_mode"]

    await interaction.response.defer(ephemeral=False)

    # 言語の確認
    file = str(interaction.guild.id) + ".ndjson"

    with open('./language_json/' + file) as f:
        read_data = ndjson.load(f)

    language = read_data[0]["language_mode"]

    # コマンドプログラム本体
    # seasonの指定がない場合は最新のシーズンの物を表示する
    if season == None:
        season = now_season

    # コースデータがあるかの確認
    files = glob.glob('./user_json/' + str(interaction.user.id)  + '/' + str(season) + "/*.ndjson")
    judge = 0

    if len(files) == 0:
        if language == "ja":
            embed=discord.Embed(title="あなたのデータはありません！", color=0xff0000)
            embed.add_field(name="順位の記録を追加してください。", value=" ", inline=False)
        elif language == "en":
            embed=discord.Embed(title="We do not have your data!", color=0xff0000)
            embed.add_field(name="Please add a record of your rank.", value=" ", inline=False)
        
        await interaction.followup.send(embed=embed)
        return

    course_name = []
    average = []
    for i in range(0, len(files)):
        #コースデータの読込
        with open('./user_json/' + str(interaction.user.id)  + '/' + str(season) + '/' + os.path.split(files[i])[1]) as f:
            read_data = ndjson.load(f)
        
        # コース名を配列に入れる
        name = Track.from_nick(os.path.splitext(os.path.split(files[i])[1])[0])
        if language == "ja":
            course_name.append(name.full_name_ja)
        elif language == "en":
            course_name.append(name.full_name)
        
        # 平均を出して配列に入れる
        # 平均を出す処理
        sum = 0
        avg = 0
        for j in range(0, len(read_data)):
            sum += read_data[j]["rank"]
        
        average.append(sum / len(read_data))
        #print(course_name[i])

    hozon_num = 0
    hozon_str = ""
    # display_modeがascだった場合
    if display_mode == "asc":
        # 大きい順に並び変える
        # （本来はascは小さい順だが、マリオカートの順位の大きいから小さい順番は1→12となるため反対の意味を表す）
        for i in range(0, len(average)):
            for j in range(0, len(average)):
                if(average[i] < average[j]):
                    hozon_num = average[i]
                    average[i] = average[j]
                    average[j] = hozon_num

                    hozon_str = course_name[i]
                    course_name[i] = course_name[j]
                    course_name[j] = hozon_str
                
    # display_modeがdescだった場合
    elif display_mode == "desc":
        # 小さい順に並び変える
        # （本来はdescは大きい順だが、マリオカートの順位の小さいから大きい順番は12→1のため反対の意味を表す）
        for i in range(0, len(average)):
            for j in range(0, len(average)):
                if(average[i] > average[j]):
                    hozon_num = average[i]
                    average[i] = average[j]
                    average[j] = hozon_num

                    hozon_str = course_name[i]
                    course_name[i] = course_name[j]
                    course_name[j] = hozon_str
    
    # 表示を行う
    cut = 25
    if language == "ja":
        embed=discord.Embed(title="Season" + str(season) + " 順位平均リスト", color=0x00008b)
    elif language == "en":
        embed=discord.Embed(title="Season" + str(season) + " Rank Average list", color=0x00008b)
    
    embed.set_author(name=interaction.user.name, icon_url=interaction.user.avatar)
    for i in range(0, len(average)):
        if language == "ja":
            embed.add_field(name="", value=str(i+1) + ". " + course_name[i] + "：" + str(round(average[i], 1)) + "位", inline=False)
        elif language == "en":
            embed.add_field(name="", value=str(i+1) + ". " + course_name[i] + "：" + str(round(average[i], 1)), inline=False)
        
        #print(i)
        #print(cut)
        if i == len(files) - 1:
            #表示させる
            await interaction.followup.send(embed=embed)
            return
        
        if i + 1 == cut:
            cut = cut + 25
            #表示させる
            await interaction.followup.send(embed=embed)
            embed=discord.Embed(title="", color=0x00008b)
            
            # DiscordのWebhook送信制限に引っかからないための対策　※効果があるかは不明
            await asyncio.sleep(2)

# /average
@tree.command(name="average",description="入力されたコースの平均順位を表示します。 / Typing course display average rank.")
@discord.app_commands.guild_only()
@discord.app_commands.choices(season=[discord.app_commands.Choice(name="9",value="9")])
async def average_command(interaction: discord.Interaction,course:str,last:int,season:str = None):

    await interaction.response.defer(ephemeral=False)

    # 言語の確認
    file = str(interaction.guild.id) + ".ndjson"

    with open('./language_json/' + file) as f:
        read_data = ndjson.load(f)

    language = read_data[0]["language_mode"]

    # コマンドプログラム本体
    # seasonの指定がない場合は最新のシーズンの物を表示する
    if season == None:
        season = now_season
    
    # コース名を表示させるための準備としてnameに入れる
    try:
        name = Track.from_nick(course)
        print(name.abbr)
    except Exception as e:
        if language == "ja":
            embed=discord.Embed(title="そのコースデータは存在しません！", color=0xff0000)
            embed.add_field(name="入力されたコース名を確認してください。", value=" ", inline=False)
        elif language == "en":
            embed=discord.Embed(title="That course data does not exist!", color=0xff0000)
            embed.add_field(name="Please confirm the course name entered.", value=" ", inline=False)
        
        await interaction.followup.send(embed=embed)
        return
    
    # コースデータがあるかの確認
    files = glob.glob('./user_json/' + str(interaction.user.id)  + '/' + str(season) + "/*.ndjson")
    judge = 0

    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if(os.path.split(files[i])[1] == name.abbr + ".ndjson"):
            judge = 1
            break
    
    if judge == 0:
        if language == "ja":
            embed=discord.Embed(title="入力されたコースデータはあなたのデータにはありません！", color=0xff0000)
            embed.add_field(name="入力されたコース名を確認してください。", value=" ", inline=False)
        elif language == "en":
            embed=discord.Embed(title="The course data entered is not in your data!", color=0xff0000)
            embed.add_field(name="Please confirm the course name entered.", value=" ", inline=False)
        
        await interaction.followup.send(embed=embed)
        return
    
    #コースデータの読込
    with open('./user_json/' + str(interaction.user.id)  + '/' + str(season) + '/' + name.abbr + ".ndjson") as f:
        read_data = ndjson.load(f)

    if len(read_data) < last:
        if language == "ja":
            embed=discord.Embed(title="指定されたコース記録の表示できる記録の数よりも数が多く指定されています！", description="このコースの記録されている順位の数は" + str(len(read_data)) + "個です。", color=0xff0000)
        elif language == "en":
            embed=discord.Embed(title="More records are specified than can be displayed for a given course record!", description="The number of recorded positions for this course is " + str(len(read_data)) + ".", color=0xff0000)
        await interaction.followup.send(embed=embed)
        return
    
    # 平均を出す処理
    sum = 0
    avg = 0
    for i in range(0, len(read_data)):
        sum += read_data[i]["rank"]
    
    avg = sum / len(read_data)

    # 表示させる
    # 平均の表示
    if language == "ja":
        embed=discord.Embed(title="Season" + str(season) + " 平均順位", color=0x00008b)
    elif language == "en":
        embed=discord.Embed(title="Season" + str(season) + " Average Rank", color=0x00008b)
    embed.set_author(name=interaction.user.name, icon_url=interaction.user.avatar)
    if language == "ja":
        embed.add_field(name=name.full_name_ja + "：" + str(round(avg, 1)) + "位", value="", inline=False)
    elif language == "en":
        embed.add_field(name=name.full_name + "：" + str(round(avg, 1)), value="", inline=False)
    # 指定された回数分の過去記録の表示
    cut = 23 # ここは23よりも大きい数は入れてはいけない（embed.add_fieldの限界個数が25個？ぽいから）
    if language == "ja":
        embed.add_field(name="過去の記録\n※直近" + str(last) + "回分", value="")
    elif language == "en":
        embed.add_field(name="Old Records\n*Last " + str(last) + " records", value="")
    
    for i in range(len(read_data)-1, len(read_data)-(1+last), -1):

        if language == "ja":
            date = dateutil.parser.parse(read_data[i]["time"]).astimezone(JST)
            embed.add_field(name="", value=date.strftime("%Y/%m/%d %H:%M:%S") + " JST : " + str(read_data[i]["rank"]) + "位", inline=False)
        elif language == "en":
            date = dateutil.parser.parse(read_data[i]["time"]).astimezone(UTC)
            embed.add_field(name="", value=date.strftime("%Y/%m/%d %H:%M:%S") + " UTC : " + str(read_data[i]["rank"]), inline=False)

        #print(i)
        #print(cut)
        if i == len(read_data)-(1+last) + 1:
            #表示させる
            embed.set_image(url="https://ay2416.github.io/Rank-Collector/stage_picture/" + str(name.id) + ".jpg")
            embed.set_footer(text="Picture : ©Mario Kart Blog")   
            await interaction.followup.send(embed=embed)
            break
        
        if i + 1 == cut:
            cut = cut + 25
            #表示させる
            await interaction.followup.send(embed=embed)
            embed=discord.Embed(title="", color=0x00008b)

            # DiscordのWebhook送信制限に引っかからないための対策　※効果があるかは不明
            await asyncio.sleep(2)

# /stage
@tree.command(name="stage",description="入力されたコースの呼び方を全て表示します。 / Typing course display all nickname.")
@discord.app_commands.guild_only()
async def stage_command(interaction: discord.Interaction,course:str):
    # 言語の確認
    file = str(interaction.guild.id) + ".ndjson"

    with open('./language_json/' + file) as f:
        read_data = ndjson.load(f)

    language = read_data[0]["language_mode"]

    # コマンドプログラム本体
    # エラーを履いたらDiscord上にエラーメッセージを出すようにする（大体コースが存在しない場合）
    try:
        name = Track.from_nick(course)
        print(name.abbr)
    except Exception as e:
        if language == "ja":
            embed=discord.Embed(title="検索結果（" + course + ")",description="このコース名ではヒットしませんでした。",color=0xff0000)
        elif language == "en":
            embed=discord.Embed(title="Search Result（" + course + ")",description="No hits were found with this course name.",color=0xff0000)
        await interaction.response.send_message(embed=embed, ephemeral=False)
        return
    
    # メッセージ表示
    if language == "ja":
        
        embed=discord.Embed(title="検索結果（" + course + ")",color=0x00ff40)
        embed.add_field(name="id", value=str(name.id), inline=False)
        embed.add_field(name="__ ・英語__", value="", inline=False)
        embed.add_field(name="略称", value=name.abbr, inline=False)
        embed.add_field(name="呼称", value=name.name, inline=False)
        embed.add_field(name="正式名称", value=name.full_name, inline=False)
        embed.add_field(name="__ ・日本語__", value="", inline=False)
        embed.add_field(name="略称", value=name.abbr_ja, inline=False)
        embed.add_field(name="呼称", value=name.name_ja, inline=False)
        embed.add_field(name="正式名称", value=name.full_name_ja, inline=False)
        embed.set_image(url="https://ay2416.github.io/Rank-Collector/stage_picture/" + str(name.id) + ".jpg")
        embed.set_footer(text="Picture : ©Mario Kart Blog") 
    elif language == "en":
        embed=discord.Embed(title="Search Result（" + course + ")",color=0x00ff40)
        embed.add_field(name="id", value=str(name.id), inline=False)
        embed.add_field(name="__ ・English__", value="", inline=False)
        embed.add_field(name="abbreviation", value=name.abbr, inline=False)
        embed.add_field(name="name", value=name.name, inline=False)
        embed.add_field(name="full name", value=name.full_name, inline=False)
        embed.add_field(name="__ ・Japanese__", value="", inline=False)
        embed.add_field(name="abbreviation", value=name.abbr_ja, inline=False)
        embed.add_field(name="name", value=name.name_ja, inline=False)
        embed.add_field(name="full name", value=name.full_name_ja, inline=False)
        embed.set_image(url="https://ay2416.github.io/Rank-Collector/stage_picture/" + str(name.id) + ".jpg")
        embed.set_footer(text="Picture : ©Mario Kart Blog") 
    
    await interaction.response.send_message(embed=embed, ephemeral=False)

# /backup
@tree.command(name="backup",description="データをバックアップします / Backup data.")
@discord.app_commands.guild_only()
async def backup_command(interaction: discord.Interaction):

    await interaction.response.defer(ephemeral=True)

    # 言語の確認
    file = str(interaction.guild.id) + ".ndjson"

    with open('./language_json/' + file) as f:
        read_data = ndjson.load(f)

    language = read_data[0]["language_mode"]

    # user_jsonにデータがあるかの確認
    files = glob.glob('./user_json/*')
    judge = 0

    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if os.path.split(files[i])[1] == str(interaction.user.id):
            judge = 1
            break 
    
    if judge == 0:
        if language == "ja":
            embed=discord.Embed(title="あなたのデータがありません！", description="データを入力してからお試しください。", color=0xff0000)
        elif language == "en":
            embed=discord.Embed(title="We do not have your data!", description="Please try it after entering your data.", color=0xff0000)
        
        await interaction.followup.send(embed=embed)
        return
    
    # データをzipに圧縮
    shutil.make_archive(str(interaction.user.id), format='zip', root_dir="./user_json/" + str(interaction.user.id))

    # Discordに送信
    if language == "ja":
        embed=discord.Embed(title="データの出力に成功しました！", description="データを改変せずに、保存しておいてください。\nもしこのデータの復元がしたい際には、\n/supportコマンドの場所までご連絡ください。", color=0x00ff40)
    elif language == "en":
        embed=discord.Embed(title="Data output succeeded!", description="Please keep the data unaltered.\nIf you would like to restore this data,\n please contact us at the /support command location.", color=0x00ff40)    
    
    await interaction.followup.send(embed=embed,file=discord.File('./' + str(interaction.user.id) + '.zip'))
    
    # zipファイルの削除
    os.remove('./' + str(interaction.user.id) + '.zip')
    #await interaction.response.send_message("コマンドがまだ実装されていません！",ephemeral=False)

# /restore
@tree.command(name="restore",description="データを復元します。 / Restore data.")
@discord.app_commands.guild_only()
async def restore_command(interaction: discord.Interaction):
    # 言語の確認
    file = str(interaction.guild.id) + ".ndjson"

    with open('./language_json/' + file) as f:
        read_data = ndjson.load(f)

    language = read_data[0]["language_mode"]

    if language == "ja":
        await interaction.response.send_message("ウイルス対策の観点からまだ実装できていません。\n申し訳ありません。\nバックアップしたファイルの復元をしたい際は、/supportコマンドの場所までご連絡ください。",ephemeral=True)
    elif language == "en":
        await interaction.response.send_message("It has not yet been implemented from an virus perspective.\nOur apologies. \nIf you would like to restore your backed up files,\nplease contact us at the /support command location.",ephemeral=True)

# /export
@tree.command(name="export",description="指定されたシーズンの保存されているデータを指定された形式で出力します。 / Typing season data export various file format.")
@discord.app_commands.guild_only()
@discord.app_commands.choices(file_format=[discord.app_commands.Choice(name="Excel",value="Excel"),discord.app_commands.Choice(name="PDF",value="PDF"),
                                           discord.app_commands.Choice(name="jpg",value="jpg")])
@discord.app_commands.choices(season=[discord.app_commands.Choice(name="9",value="9")])
async def exoprt_command(interaction: discord.Interaction,file_format:str,season:str = None):

    await interaction.response.defer(ephemeral=True)

    # 言語の確認
    file = str(interaction.guild.id) + ".ndjson"

    with open('./language_json/' + file) as f:
        read_data = ndjson.load(f)

    language = read_data[0]["language_mode"]

    # seasonの指定がない場合は最新のシーズンの物を表示する
    if season == None:
        season = now_season
    
    # user_jsonにデータがあるかの確認
    files = glob.glob('./user_json/*')
    judge = 0

    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if os.path.split(files[i])[1] == str(interaction.user.id):
            judge = 1
            break 
    
    if judge == 0:
        if language == "ja":
            embed=discord.Embed(title="あなたのデータがありません！", description="データを入力してからお試しください。", color=0xff0000)
        elif language == "en":
            embed=discord.Embed(title="We do not have your data!", description="Please try it after entering your data.", color=0xff0000)    
        await interaction.followup.send(embed=embed)
        return
       
    # 指定されたファイル形式によって処理を分ける
    if file_format == "Excel":
        # ベースファイルをコピー
        shutil.copyfile("base.xlsx", str(interaction.user.id) + ".xlsx")

        # データを入れる
        book = openpyxl.load_workbook('./' + str(interaction.user.id) + '.xlsx')
        sheet = book['Record Write']

        # 名前、シーズン数、時間の挿入
        sheet.cell(row=2, column=3).value = str(interaction.user.name)
        sheet.cell(row=3, column=3).value = season

        if language == "ja":
            date = dateutil.parser.parse(str(datetime.datetime.now())).astimezone(JST)
            sheet.cell(row=4, column=3).value = date.strftime("%Y/%m/%d %H:%M:%S") + " JST"
        elif language == "en":
            date = dateutil.parser.parse(str(datetime.datetime.now())).astimezone(UTC)
            sheet.cell(row=4, column=3).value = date.strftime("%Y/%m/%d %H:%M:%S") + " UTC"
        
        for i in range(0, 96):
            # セルからコースの略称を取得
            # 正式名称を入れていく
            if sheet.cell(row=7, column=4+i).value == None:
                sheet.cell(row=8, column=4+i).value = "No data."
            else:
                course = sheet.cell(row=7, column=4+i).value
                #print(course)
                name = Track.from_nick(course)

                if language == "ja":
                    sheet.cell(row=8, column=4+i).value = name.full_name_ja
                elif language == "en":
                    sheet.cell(row=8, column=4+i).value = name.full_name
        
        # コースデータにあった順位データを入れていく
        for i in range(0, 96):

            if sheet.cell(row=7, column=4+i).value == None:
                continue
            else:
                course = sheet.cell(row=7, column=4+i).value
                #print(course)
                name = Track.from_nick(course)
                #print(name.abbr)
                # コースデータがあるかの確認
                files = glob.glob('./user_json/' + str(interaction.user.id)  + '/' + str(season) + "/*.ndjson")
                judge = 0

                for j in range(0, len(files)):
                    #print(os.path.split(files[i])[1])
                    if(os.path.split(files[j])[1] == name.abbr + ".ndjson"):
                        judge = 1
                        break
                
                if judge == 1:
                    #コースデータの読込
                    with open('./user_json/' + str(interaction.user.id)  + '/' + str(season) + '/' + name.abbr + ".ndjson") as f:
                        read_data = ndjson.load(f)

                    for k in range(0, len(read_data)):
                        sheet.cell(row=9+k, column=4+i).value = read_data[k]["rank"]
                        #print(sheet.cell(row=9+k, column=4+i).value)

        #　保存を行う
        book.save('./' + str(interaction.user.id) + '.xlsx')

        # Discord上に送信
        if language == "ja":
            embed=discord.Embed(title="データの出力に成功しました！", description="何かのお役に立てると幸いです！", color=0x00ff40)
        elif language == "en":
            embed=discord.Embed(title="Data output succeeded!", description="I hope I can be of some help!", color=0x00ff40)
        
        await interaction.followup.send(embed=embed,file=discord.File('./' + str(interaction.user.id) + '.xlsx'))
        
        # xlsxファイルの削除
        os.remove('./' + str(interaction.user.id) + '.xlsx')
    # PDFが指定された場合　※エクセルシートをPDFに変換して行う
    elif file_format == "PDF":
        await interaction.followup.send("まだ実装されていません！")
        '''
        # ベースファイルをコピー
        shutil.copyfile("base.xlsx", str(interaction.user.id) + ".xlsx")

        # データを入れる
        book = openpyxl.load_workbook('./' + str(interaction.user.id) + '.xlsx')
        sheet = book['Record Write']

        # 名前、シーズン数、時間の挿入
        sheet.cell(row=2, column=3).value = str(interaction.user.name)
        sheet.cell(row=3, column=3).value = season

        if language == "ja":
            date = dateutil.parser.parse(str(datetime.datetime.now())).astimezone(JST)
            sheet.cell(row=4, column=3).value = date.strftime("%Y/%m/%d %H:%M:%S") + " JST"
        elif language == "en":
            date = dateutil.parser.parse(str(datetime.datetime.now())).astimezone(UTC)
            sheet.cell(row=4, column=3).value = date.strftime("%Y/%m/%d %H:%M:%S") + " UTC"
        
        for i in range(0, 96):
            # セルからコースの略称を取得
            # 正式名称を入れていく
            if sheet.cell(row=7, column=4+i).value == None:
                sheet.cell(row=8, column=4+i).value = "No data."
            else:
                course = sheet.cell(row=7, column=4+i).value
                #print(course)
                name = Track.from_nick(course)

                if language == "ja":
                    sheet.cell(row=8, column=4+i).value = name.full_name_ja
                elif language == "en":
                    sheet.cell(row=8, column=4+i).value = name.full_name
        
        # コースデータにあった順位データを入れていく
        for i in range(0, 96):

            if sheet.cell(row=7, column=4+i).value == None:
                continue
            else:
                course = sheet.cell(row=7, column=4+i).value
                #print(course)
                name = Track.from_nick(course)
                #print(name.abbr)
                # コースデータがあるかの確認
                files = glob.glob('./user_json/' + str(interaction.user.id)  + '/' + str(season) + "/*.ndjson")
                judge = 0

                for j in range(0, len(files)):
                    #print(os.path.split(files[i])[1])
                    if(os.path.split(files[j])[1] == name.abbr + ".ndjson"):
                        judge = 1
                        break
                
                if judge == 1:
                    #コースデータの読込
                    with open('./user_json/' + str(interaction.user.id)  + '/' + str(season) + '/' + name.abbr + ".ndjson") as f:
                        read_data = ndjson.load(f)

                    for k in range(0, len(read_data)):
                        sheet.cell(row=9+k, column=4+i).value = read_data[k]["rank"]
                        #print(sheet.cell(row=9+k, column=4+i).value)

        # 保存を行う
        book.save('./' + str(interaction.user.id) + '.xlsx')

        # excelファイルをpdfファイルへ

        # Discord上に送信
        if language == "ja":
            embed=discord.Embed(title="データの出力に成功しました！", description="何かのお役に立てると幸いです！", color=0x00ff40)
        elif language == "en":
            embed=discord.Embed(title="Data output succeeded!", description="I hope I can be of some help!", color=0x00ff40)
        
        await interaction.followup.send(embed=embed,file=discord.File('./' + str(interaction.user.id) + '.pdf'))
        
        # xlsx, pdfファイルの削除
        os.remove('./' + str(interaction.user.id) + '.xlsx')
        os.remove('./' + str(interaction.user.id) + '.pdf')
        '''
    # jpgだった場合
    elif file_format == "jpg":
        await interaction.followup.send("まだ実装されていません！")
    #await interaction.response.send_message("コマンドがまだ実装されていません！",ephemeral=False)

# /language
@tree.command(name="language",description="言語を変更します。（jaまたはen） / Change language. (ja or en)")
@discord.app_commands.guild_only()
@discord.app_commands.choices(language=[discord.app_commands.Choice(name="ja",value="ja"),discord.app_commands.Choice(name="en",value="en")])
async def language_command(interaction: discord.Interaction,language:str):

    # 言語の確認
    file = str(interaction.guild.id) + ".ndjson"

    with open('./language_json/' + file) as f:
        read_data = ndjson.load(f)

    now_language = read_data[0]["language_mode"]

    # コマンドプログラム本体
    # 登録されている言語かどうかの確認
    if language == "ja" or language == "en":
        print("成功！:登録されている言語です！")
    else:
        print("エラー！:コマンドの言語の設定が間違っています！ご確認ください。")
        if now_language == "ja":
            embed=discord.Embed(title="エラー！", description="コマンドの言語指定が間違っています！\nご確認ください。", color=0xff0000)
        elif now_language == "en":
            embed=discord.Embed(title="Error!", description="Command language setting error!\nCheck typing keyword. ([ja] or [en])", color=0xff0000)
        await interaction.response.send_message(embed=embed)
        return
    
    # 既にファイルが存在しているかの判定
    files = glob.glob('./language_json/*.ndjson')
    judge = 0

    for i in range(0, len(files)):
        #print(os.path.split(files[i])[1])
        if(os.path.split(files[i])[1] == str(interaction.guild.id) + ".ndjson"):
            print("一致しました！")
            judge = 1
            break
        else:
            judge = 0
    
    file = str(interaction.guild.id) + ".ndjson"

    if(judge == 1):
        os.remove("./language_json/" + file)

    content = {
        "language_mode" : language
    }

    with open('./language_json/' + file, 'a') as f:
        writer = ndjson.writer(f)
        writer.writerow(content)

    # メッセージ表示
    if language == "ja":
        print(str(interaction.guild.id) + "の言語を日本語に変更しました。")
        embed=discord.Embed(title="成功しました!", description="日本語に変更しました。", color=0x00ff40)
        await interaction.response.send_message(embed=embed, ephemeral=False)
    elif language == "en":
        print(str(interaction.guild.id) + "の言語を英語に変更しました。")
        embed=discord.Embed(title="Success!", description="Change to English.", color=0x00ff40)
        await interaction.response.send_message(embed=embed, ephemeral=False)

# /help
@tree.command(name="help",description="コマンドについての簡単な使い方を出します。 / How to use command and Command list.")
@discord.app_commands.guild_only()
async def help_command(interaction: discord.Interaction):
    # 言語の確認
    file = str(interaction.guild.id) + ".ndjson"

    with open('./language_json/' + file) as f:
        read_data = ndjson.load(f)

    language = read_data[0]["language_mode"]

    # コマンドプログラム本体
    #Discord上にヘルプを表示
    if language == "ja":
        embed=discord.Embed(title="コマンドリスト")
        embed.add_field(name="/add mode:[start / stop]", value="順位の記録の追加を開始・停止します。\n※10分間操作がなければ自動で停止します。", inline=False)
        embed.add_field(name="/delete mode:[start / stop]", value="順位の記録の削除を開始・停止します。\n※10分間操作がなければ自動で停止します。", inline=False)
        embed.add_field(name="/list_avg display_mode:[降順(12→1) / 昇順(1→12)] season:[シーズン数]", value="登録されている全てのコースの平均順位を表示します。\n※「season」の指定がなければ、最新のシーズンの保存してある記録を表示します。", inline=False)
        embed.add_field(name="/average course:[コース名] last:[直近何回分を表示するか] season:[シーズン数]", value="指定されたコースの平均順位を表示します。\n※「last」は過去の記録を表示する数です。\n※「season」の指定がなければ、最新のシーズンの保存してある記録を表示します。", inline=False)
        embed.add_field(name="/stage course:[コース名]", value="入力されたコースの呼び名を全て表示します。（登録されているもののみ）", inline=False)
        embed.add_field(name="/backup", value="このBotに保存されている順位データをバックアップし、ダウンロードをできるようにします。", inline=False)
        embed.add_field(name="/restore", value="順位データを復元します。\n※このコマンドはまだ実装できていません。\n復元を希望の場合は、/supportコマンドの場所までご連絡をください。", inline=False)
        embed.add_field(name="/export file_format:[エクスポートする形式]", value="このBotに保存されている順位データを指定の形式にエクスポートします。", inline=False)
        embed.add_field(name="/language language:[言語の選択（ja/en）]", value="このBotのコマンドの言語を変更します。", inline=False)
        embed.add_field(name="/help", value="このBotのコマンドの簡単な使い方を出します。", inline=False)
    elif language == "en":
        embed=discord.Embed(title="Command list")
        embed.add_field(name="/add mode:[start / stop]", value="Starts and stops the addition of rank records.\n*If there is no operation for 10 minutes, it will automatically stop.", inline=False)
        embed.add_field(name="/delete mode:[start / stop]", value="Starts and stops the deletion of rank records.\n*If there is no operation for 10 minutes, it will automatically stop.", inline=False)
        embed.add_field(name="/list_avg display_mode:[desc(12→1) / asc(1→12)] season:[season num]", value="Displays the average ranking of all registered courses.\n*If [season] is not specified, the stored records for the most recent season are displayed.", inline=False)
        embed.add_field(name="/average course:[course name] last:[last records display num] season:[season num]", value="Displays the average ranking for a given course.\n*[Last] is the number of past records to be displayed.\n*If [season] is not specified, the stored records for the most recent season are displayed.", inline=False)
        embed.add_field(name="/stage course:[course name]", value="Displays all course call names entered. (Only those that have been registered)", inline=False)
        embed.add_field(name="/backup", value="Backup the ranking data stored in this bot and make it available for download.", inline=False)
        embed.add_field(name="/restore", value="Restore rank data. \n*This command has not yet been implemented.\nIf you wish to restore, please contact us at the /support command location.", inline=False)
        embed.add_field(name="/export file_format:[export format]", value="Export the ranking data stored in this bot to the specified format.", inline=False)
        embed.add_field(name="/language language:[choose language（ja/en）]", value="Change default language.", inline=False)
        embed.add_field(name="/help", value="How to use command and Command list.", inline=False)
    '''
    if interaction.guild.id == your_guild_id: # もしサーバー限定コマンドの実装があった場合の表記をした場合はここに書く
        embed.add_field(name="", value="", inline=False)
    '''
    await interaction.response.send_message(embed=embed,ephemeral=False)

client.run(os.environ['token'])